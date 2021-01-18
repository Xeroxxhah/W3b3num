import socket
import sys
import openpyxl as xl
import urllib.request
import argparse
import requests
import nmap
import os

argparse = argparse.ArgumentParser(description="Web Enum Tool")
argparse.add_argument('host', type=str, help='Host Address (www.example.com)')
args = argparse.parse_args()
# CODE Function
try:
    h = args.host
    host = h.split('.')
    hostdir = host[1]
    os.mkdir(hostdir)
except FileExistsError:
    pass

subdomains_list = []
subdomains_IP_list = []
twohundred_list = []
simpleScan = True
intenseScan = False

if simpleScan:
    ports = 1000
elif intenseScan:
    ports = 65535

def seprateProtocol(url):
    if 'https' == url[:5].lower():
        return url[:5]
    elif 'http' == url[:4]:
        return url[:4]
    else:
        print('Wrong url format\nformat: https://www.example.com')
        quit()



def CODE(URL):
    try:
       xurl = urllib.request.urlopen(URL)
       code = xurl.getcode()
       return code
    except ValueError:
        print("URL_FORMAT_ERROR:\nformat:https://www.google.com")
        quit()
    except urllib.error.HTTPError:
        return 404
    except urllib.error.URLError:
        print('ERROR:NO INTERNET CONNECTION(-*-)')
        quit()



def subDomainEnum():
    protocol = seprateProtocol(args.host)
    h = args.host
    host = h.split('.')
    if len(host) > 3:
        formated_host = host[1]+'.'+host[2]+'.'+host[3]
    else:
        formated_host = host[1]+'.'+host[2]
    subdomain_file = open('subdomains.txt')
    content = subdomain_file.read()
    subdomains = content.splitlines()
    for subdomain in subdomains:
        URL = f'{protocol}://{subdomain}.{formated_host}'
        try:
            requests.get(URL)
        except requests.ConnectionError:
            pass
        else:
            print('[+] ',URL)
            subdomains_list.append(URL)


def stripHost(host):
    if  'https' == host[:5].lower():
        return host[8:]
    else:
        return host[7:]


def getip():
    for host in subdomains_list:
        IP = socket.gethostbyname(stripHost(host))
        subdomains_IP_list.append(IP)

def countLines():
    try:
        with open('list.txt') as f:
            return len(f.readlines())
    except FileExistsError:
        pass

def get200(host):
    print(f'There are {countLines()} words.\n to exit this test press ctrl + c')
    f = open("list.txt")
    for word in f:
        try:
            print(host+'/'+word ,CODE(host+'/'+word))
            if CODE(host+'/'+word) == 200:
                twohundred_list.append(host+'/'+word)
        except KeyboardInterrupt:
            break
                



def Xlwork():
    workbook = xl.load_workbook('sample.xlsx')
    sheet = workbook['Sheet1']
    cell = sheet[f'a1']
    cell.value = 'Sudomains'
    cell = sheet[f'b1']
    cell.value = 'IPs'
    cell = sheet[f'c1']
    cell.value = 'Directory Enum'
    try:
        for i in range(2,len(subdomains_list)):
            cell = sheet[f'a{i}']
            cell.value = subdomains_list[i]
        cell = sheet[f'a{sheet.max_row+1}']
        cell.value = subdomains_list[0]
    except IndexError:
        pass
    getip()

    try:
        for i in range(2,len(subdomains_IP_list)):
            cell2 = sheet[f'b{i}']
            cell2.value = subdomains_IP_list[i]
        cell = sheet[f'b{sheet.max_row}']
        cell.value = subdomains_IP_list[0]
    except IndexError:
        pass
    try:
        for i in range(2,len(twohundred_list)):
            cell2 = sheet[f'c{i}']
            cell2.value = twohundred_list[i]
        cell = sheet[f'c{sheet.max_row+1}']
        cell.value = twohundred_list[0]
    except IndexError:
        pass
    
    workbook.save('./'+hostdir+'/'+'sample2.xlsx')


def NMAP(host):
    ip_host = socket.gethostbyname(stripHost(host))
    try:
        print('\n*NMAP SCAN*\n')
        nm = nmap.PortScanner()
        nm.scan(ip_host, '1-1024', '-v -sS',sudo=True)
        print('VERSION: ',nm.nmap_version())
        print(nm.scaninfo())
        print('Status: ',nm[ip_host].state())
        print('PORTS: ', nm[ip_host]['tcp'].keys())
        n_ver = str(nm.nmap_version())
        n_sinfo = str(nm.scaninfo())
        n_state = str(nm[ip_host].state())
        n_ports = str(nm[ip_host]['tcp'].keys())
        with open('./'+hostdir+'/'+'nmap.txt','a') as f:
            f.write('\n\n'+host+'\n'+n_ver+'\n'+n_sinfo+'\n'+n_state+'\n'+n_ports+'\n')
    except:
        pass
    


def subGet200(host):
    print(f'There are {countLines()} words.\n to exit this test press ctrl + c')
    f = open("list.txt")
    for word in f:
        try:
            print(host+'/'+word ,CODE(host+'/'+word))
            if CODE(host+'/'+word) == 200:
                with open('./'+hostdir+'/'+'Subdomain_DirEnum.txt','a') as f:
                    f.write(host+'/'+word)
        except KeyboardInterrupt:
            break



def MAIN():
    subdomain_test = False
    dirEnum_test = False
    subdomain_DirEnum = False
    NmapTest = False
    SubNmapTest = False
    subdomain = input('Do you want to run subdomain test (y/n): ')
    if subdomain.lower() == 'y':
        subdomain_test = True
    dirEnum = input('Do you want to run Directory Enumration test (y/n): ')
    if dirEnum.lower() == 'y':
        dirEnum_test = True
    subdomain = input('Do you want to run subdomain_DirEnum test (y/n) \n(This test might take a long time): ')
    if subdomain.lower() == 'y':
        subdomain_DirEnum = True
    nMAP = input('Do you want to run NMAP (y/n): ')
    if nMAP.lower() == 'y':
        NmapTest = True
    SubnMAP = input('Do you want to run NMAP on subdomain (y/n) \n(This test might take a long time): ')
    if SubnMAP.lower() == 'y':
        SubNmapTest = True
    ###############################################################
    if subdomain_test == True:
        print('-------------------------------Subdomain-Test---------------------------')
        subDomainEnum()
    if dirEnum_test == True:
        print('-------------------------------DirectoryEnum-Test---------------------------')
        get200(args.host)
    if subdomain_DirEnum == True:
        print('-------------------------------subdomain Directory Enum Test---------------------------')
        print('This test may take a long time.')
        for sub in subdomains_list:
            subGet200(sub)
    if NmapTest == True:
        print('-------------------------------NMAP---------------------------')
        NMAP(args.host)
    if SubNmapTest == True:
        print('-------------------------------Running Nmap on subdomains---------------------------')
        for sub in subdomains_list:
            NMAP(sub)
    
    Xlwork()
   


MAIN()